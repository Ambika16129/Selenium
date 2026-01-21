import  openpyxl
def row_count(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    return sheet.max_row

def column_count(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    return sheet.max_column

def write_data(path,sheetname,row_no,col_no,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    sheet.cell(row_no,col_no).value=data
    workbook.save(path)

def read_data(path,sheetname,row_no,col_no):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.cell(row_no,col_no).value

def read_login_data(path):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook['Sheet1']
    rows=sheet.max_row
    data=[]
    for r in range(2,rows+1):
        username=sheet.cell(r,1).value
        password=sheet.cell(r,2).value
        data.append((username,password))
    return data